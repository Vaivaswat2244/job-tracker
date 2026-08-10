"""One-way export: SQLite -> tracker.xlsx. State is never read back from the sheet."""
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

STATUS_FILL = {
    "found": "D9D9D9",       # grey
    "applied": "FFE699",     # yellow
    "followed_up": "F8CBAD", # orange
    "in_process": "BDD7EE",  # blue
    "offer": "C6E0B4",       # green
    "rejected": "F4B183",    # red
    "ghosted": "F4B183",     # red
}

HEADERS = [
    "app_id", "job_id", "company", "title", "status", "applied_at",
    "followup_due", "days_since_applied", "source", "comp_model",
    "pay_min", "pay_max", "currency", "auth_required", "contacts", "url", "notes",
]

QUERY = """
SELECT a.id AS app_id, j.id AS job_id, c.name AS company, j.title, a.status,
       a.applied_at, a.followup_due, j.source, j.comp_model, j.pay_min, j.pay_max,
       j.pay_currency, j.auth_required, j.url, a.notes,
       (SELECT count(*) FROM contacts ct WHERE ct.company_id = c.id) AS contacts,
       CAST(julianday('now') - julianday(a.applied_at) AS INTEGER) AS days_since_applied
FROM applications a
JOIN jobs j ON j.id = a.job_id
JOIN companies c ON c.id = j.company_id
ORDER BY (a.followup_due IS NULL), a.followup_due, a.id
"""


def export(conn, path: str = "tracker.xlsx") -> tuple[str, int]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"
    ws.append(HEADERS)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(vertical="center")

    rows = conn.execute(QUERY).fetchall()
    status_col = HEADERS.index("status") + 1
    for r in rows:
        ws.append([r[h] if h in r.keys() else None for h in HEADERS])
        fill = STATUS_FILL.get(r["status"])
        if fill:
            ws.cell(row=ws.max_row, column=status_col).fill = PatternFill(
                "solid", start_color=fill, end_color=fill
            )

    widths = {"company": 22, "title": 42, "status": 13, "url": 55, "notes": 30}
    for i, h in enumerate(HEADERS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = widths.get(h, 14)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    _contacts_sheet(wb, conn)
    wb.save(path)
    return os.path.abspath(path), len(rows)


def _contacts_sheet(wb, conn) -> None:
    ws = wb.create_sheet("Contacts")
    headers = ["id", "company", "name", "title", "email", "confidence", "linkedin", "source"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for r in conn.execute(
        "SELECT ct.id, c.name AS company, ct.name, ct.title, ct.email, ct.email_confidence,"
        " ct.linkedin_url, ct.source FROM contacts ct"
        " JOIN companies c ON c.id = ct.company_id ORDER BY c.name, ct.name"
    ):
        email = r["email"]
        # INV-2: an inferred address must never be pasteable as-is.
        if email and r["email_confidence"] == "inferred":
            email = f"[UNVERIFIED: {email}]"
        ws.append([r["id"], r["company"], r["name"], r["title"], email,
                   r["email_confidence"], r["linkedin_url"], r["source"]])
    for i, w in enumerate([6, 22, 22, 26, 34, 12, 34, 16], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
